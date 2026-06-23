"""Statistics dialog for Supervertaler Workbench.

Lets the translator analyse the current project's segments against one or more
translation memories and see a Trados/memoQ-style match breakdown:
Repetitions, 101%, 100%, 95-99%, 85-94%, 75-84%, 50-74%, No match.

The analysis runs in a background QThread (StatisticsWorker); results stream
in per-TM as each completes. Results can be exported to HTML.
"""
from __future__ import annotations

import html
from typing import Dict, List

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QComboBox, QDialog, QFileDialog, QFrame, QHBoxLayout, QHeaderView, QLabel,
    QMessageBox, QProgressBar, QPushButton, QScrollArea, QTableWidget,
    QTableWidgetItem, QVBoxLayout, QWidget,
)

from modules.statistics_analyzer import (
    CATEGORIES, CATEGORY_HELP, StatisticsWorker, TMResult,
)
from modules.styled_widgets import CheckmarkCheckBox
from modules.help_system import Topics as HelpTopics, open_help, set_topic as set_help_topic


COLUMNS = ["Type", "Segments", "Words", "Characters", "Tags", "Percent"]


class StatisticsDialog(QDialog):
    """Analyse the open project against TMs and display a match breakdown."""

    def __init__(self, parent, db_manager, segments, source_lang, target_lang,
                 tm_choices: List[dict], preselected_tm_ids: List[str],
                 project_name: str = ""):
        """
        Args:
            db_manager: the app's DatabaseManager (for db_path)
            segments: list of Segment objects from the open project
            source_lang / target_lang: project language codes
            tm_choices: [{'tm_id', 'name', 'entry_count'}, ...]
            preselected_tm_ids: TM ids ticked by default (project's active TMs)
            project_name: the open project's name (shown in the header and exports)
        """
        super().__init__(parent)
        self.setWindowTitle("Statistics – Analyse Against TM")
        self.resize(900, 560)

        self._db_manager   = db_manager
        self._segments     = segments
        self._src_lang     = source_lang
        self._tgt_lang     = target_lang
        self._tm_choices   = tm_choices
        self._project_name = project_name or ""
        self._worker       = None
        self._results: List[TMResult] = []
        self._tm_checks: Dict[str, CheckmarkCheckBox] = {}

        layout = QVBoxLayout(self)

        # ---- Top bar: contextual help "?" ----------------------------
        help_row = QHBoxLayout()
        help_row.addStretch(1)
        help_btn = QPushButton("?")
        help_btn.setFixedSize(22, 22)
        help_btn.setToolTip("Open Statistics help (online)")
        help_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        help_btn.clicked.connect(lambda: open_help(HelpTopics.TOOL_STATISTICS))
        help_row.addWidget(help_btn)
        layout.addLayout(help_row)

        # ---- Header / scope summary ----------------------------------
        word_total = sum(len((s.source or "").split()) for s in segments)
        proj = f"<b>{self._project_name}</b> — " if self._project_name else ""
        hdr = QLabel(
            f"{proj}<b>{len(segments):,}</b> segments · approx. "
            f"<b>{word_total:,}</b> source words  ·  {source_lang} → {target_lang}"
        )
        hdr.setTextFormat(Qt.TextFormat.RichText)
        hdr.setWordWrap(True)
        layout.addWidget(hdr)

        # ---- TM selection --------------------------------------------
        tm_label = QLabel("Analyse against these translation memories:")
        tm_label.setStyleSheet("font-weight: bold; margin-top: 6px;")
        layout.addWidget(tm_label)

        hint = QLabel("Leave all unticked for a word count + internal repetitions only (no TM).")
        hint.setStyleSheet("color: #777; font-size: 11px;")
        layout.addWidget(hint)

        tm_box = QScrollArea()
        tm_box.setWidgetResizable(True)
        tm_box.setFrameShape(QFrame.Shape.StyledPanel)
        tm_inner = QWidget()
        tm_inner_layout = QVBoxLayout(tm_inner)
        tm_inner_layout.setContentsMargins(8, 4, 8, 4)

        if not tm_choices:
            tm_inner_layout.addWidget(QLabel("No translation memories found in the database."))
        for tm in tm_choices:
            tm_id = tm["tm_id"]
            cnt   = tm.get("entry_count", 0)
            cb = CheckmarkCheckBox(f"{tm['name']}   ({cnt:,} entries)")
            cb.setChecked(tm_id in preselected_tm_ids)
            self._tm_checks[tm_id] = cb
            tm_inner_layout.addWidget(cb)
        tm_inner_layout.addStretch(1)
        tm_box.setWidget(tm_inner)
        # Stretch=1 so the TM list fills the dialog before any analysis is run
        # (the results area below is hidden until Analyse is clicked).
        layout.addWidget(tm_box, 1)

        # ---- Matching depth ------------------------------------------
        # Lets the user trade thoroughness for speed. The fuzzy pass is the
        # expensive part on a large TM; "Standard" (75% floor) is much faster
        # than "Thorough" (50%), and "Exact only" skips fuzzy altogether.
        depth_row = QHBoxLayout()
        depth_row.addWidget(QLabel("Matching depth:"))
        self.cmb_depth = QComboBox()
        # (label, fuzzy_threshold, skip_fuzzy)
        self._depth_options = [
            ("Standard – exact + fuzzy down to 75% (faster)", 0.75, False),
            ("Thorough – exact + fuzzy down to 50% (slower)", 0.50, False),
            ("Exact matches only – skip fuzzy (fastest)",      1.00, True),
        ]
        for label, _thr, _skip in self._depth_options:
            self.cmb_depth.addItem(label)
        self.cmb_depth.setCurrentIndex(0)
        depth_row.addWidget(self.cmb_depth, 1)
        layout.addLayout(depth_row)

        # ---- Action buttons ------------------------------------------
        btn_row = QHBoxLayout()
        self.btn_analyse = QPushButton("📊 Analyse")
        self.btn_analyse.clicked.connect(self._start_analysis)
        btn_row.addWidget(self.btn_analyse)

        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.clicked.connect(self._cancel_analysis)
        btn_row.addWidget(self.btn_cancel)

        self.btn_export = QPushButton("Export…")
        self.btn_export.setToolTip("Export the results to HTML, Excel, or CSV")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self._export)
        btn_row.addWidget(self.btn_export)

        btn_row.addStretch(1)
        self.btn_close = QPushButton("Close")
        self.btn_close.clicked.connect(self.accept)
        btn_row.addWidget(self.btn_close)
        layout.addLayout(btn_row)

        # ---- Progress ------------------------------------------------
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        self.status_lbl = QLabel("")
        self.status_lbl.setStyleSheet("color: #555;")
        layout.addWidget(self.status_lbl)

        # ---- Results area --------------------------------------------
        # Hidden until Analyse is clicked, so the dialog stays compact (the TM
        # list fills the space) instead of showing a large empty panel.
        self.results_area = QScrollArea()
        self.results_area.setWidgetResizable(True)
        self.results_container = QWidget()
        self.results_layout = QVBoxLayout(self.results_container)
        self.results_layout.addStretch(1)
        self.results_area.setWidget(self.results_container)
        self.results_area.setVisible(False)
        layout.addWidget(self.results_area, 3)

        # F1 anywhere in the dialog opens the same Statistics help page.
        set_help_topic(self, HelpTopics.TOOL_STATISTICS)

    # ------------------------------------------------------------------
    # Analysis lifecycle
    # ------------------------------------------------------------------

    def _selected_tm_ids(self) -> List[str]:
        return [tm_id for tm_id, cb in self._tm_checks.items() if cb.isChecked()]

    def _start_analysis(self):
        tm_ids = self._selected_tm_ids()
        # No TM selected is allowed: the worker then reports word counts and
        # internal repetitions only (matches the "analyse without a TM" use case).
        if not self._segments:
            QMessageBox.information(self, "Statistics", "The project has no segments to analyse.")
            return

        # Clear previous results
        self._clear_results()
        self._results = []

        tm_names = {tm["tm_id"]: tm["name"] for tm in self._tm_choices}

        db_path = getattr(self._db_manager, "db_path", None)
        if not db_path:
            QMessageBox.critical(self, "Statistics", "Could not locate the TM database path.")
            return

        _label, fuzzy_threshold, skip_fuzzy = self._depth_options[self.cmb_depth.currentIndex()]

        self._worker = StatisticsWorker(
            db_path=db_path,
            segments=self._segments,
            tm_ids=tm_ids,
            tm_names=tm_names,
            source_lang=self._src_lang,
            target_lang=self._tgt_lang,
            fuzzy_threshold=fuzzy_threshold,
            skip_fuzzy=skip_fuzzy,
        )
        self._worker.progress.connect(self._on_progress)
        self._worker.tm_result.connect(self._on_tm_result)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)

        self.results_area.setVisible(True)
        self.btn_analyse.setEnabled(False)
        self.btn_export.setEnabled(False)
        self.btn_cancel.setEnabled(True)
        self.cmb_depth.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)  # busy until the worker reports progress
        self.progress.setValue(0)
        self.status_lbl.setText("Starting analysis…")
        self._worker.start()

    def _cancel_analysis(self):
        if self._worker:
            self._worker.cancel()
            self.status_lbl.setText("Cancelling…")

    def _on_progress(self, current: int, total: int, message: str):
        self.progress.setRange(0, total)
        self.progress.setValue(current)
        self.status_lbl.setText(message)

    def _on_tm_result(self, result: TMResult):
        self._results.append(result)
        self._add_result_table(result)

    def _on_finished(self, results):
        self.progress.setVisible(False)
        self.status_lbl.setText(f"Done. Analysed against {len(self._results)} TM(s).")
        self.btn_analyse.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.cmb_depth.setEnabled(True)
        self.btn_export.setEnabled(bool(self._results))
        if self._results:
            self._add_legend()
        self._worker = None

    def _add_legend(self):
        """Append a compact match-type legend below the result tables."""
        rows = "".join(
            f"<tr><td style='padding-right:8px;white-space:nowrap;vertical-align:top'>"
            f"<b>{html.escape(label)}</b></td>"
            f"<td style='color:#444'>{html.escape(CATEGORY_HELP.get(label, ''))}</td></tr>"
            for label in CATEGORIES
        )
        legend = QLabel(
            "<div style='margin-top:8px'><b>What the match types mean</b>"
            f"<table style='font-size:11px;margin-top:4px'>{rows}</table></div>"
        )
        legend.setTextFormat(Qt.TextFormat.RichText)
        legend.setWordWrap(True)
        self.results_layout.insertWidget(self.results_layout.count() - 1, legend)

    def _on_error(self, msg: str):
        self.progress.setVisible(False)
        self.btn_analyse.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.cmb_depth.setEnabled(True)
        self._worker = None
        QMessageBox.critical(self, "Statistics – Error", msg)

    # ------------------------------------------------------------------
    # Result rendering
    # ------------------------------------------------------------------

    def _clear_results(self):
        while self.results_layout.count() > 1:  # keep trailing stretch
            item = self.results_layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()

    def _add_result_table(self, result: TMResult):
        title = QLabel(f"Analysis – Resources: <b>{html.escape(result.tm_name)}</b>")
        title.setTextFormat(Qt.TextFormat.RichText)
        title.setStyleSheet("margin-top: 10px; margin-bottom: 2px;")
        self.results_layout.insertWidget(self.results_layout.count() - 1, title)

        rows = ["All"] + list(CATEGORIES)
        table = QTableWidget(len(rows), len(COLUMNS))
        table.setHorizontalHeaderLabels(COLUMNS)
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)

        total_words = result.total.words or 1

        def _bucket_for(label):
            if label == "All":
                return result.total
            return result.buckets[label]

        for r, label in enumerate(rows):
            b = _bucket_for(label)
            pct = (b.words / total_words * 100.0) if label != "All" else 100.0
            cells = [
                label,
                f"{b.segments:,}",
                f"{b.words:,}",
                f"{b.chars:,}",
                f"{b.tags:,}",
                f"{pct:.2f}",
            ]
            for c, val in enumerate(cells):
                item = QTableWidgetItem(str(val))
                if c > 0:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if label == "All":
                    f = item.font()
                    f.setBold(True)
                    item.setFont(f)
                table.setItem(r, c, item)

        hh = table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for c in range(1, len(COLUMNS)):
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.ResizeToContents)

        # Size table to fit its content (no inner scrollbar)
        table.setMinimumHeight(table.verticalHeader().length() + table.horizontalHeader().height() + 4)
        table.setMaximumHeight(table.verticalHeader().length() + table.horizontalHeader().height() + 4)

        self.results_layout.insertWidget(self.results_layout.count() - 1, table)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def _export(self):
        """Export the results, letting the user pick HTML, Excel, or CSV."""
        if not self._results:
            return
        html_flt = "Web page (*.html)"
        xlsx_flt = "Excel workbook (*.xlsx)"
        csv_flt  = "CSV file (*.csv)"
        path, selected = QFileDialog.getSaveFileName(
            self, "Export statistics", "supervertaler-statistics.html",
            ";;".join([html_flt, xlsx_flt, csv_flt]))
        if not path:
            return

        # Decide format from the chosen filter, falling back to the extension.
        lower = path.lower()
        if selected == xlsx_flt or lower.endswith(".xlsx"):
            fmt = "xlsx"
        elif selected == csv_flt or lower.endswith(".csv"):
            fmt = "csv"
        else:
            fmt = "html"
        # Ensure the path carries the right extension.
        ext = "." + fmt
        if not lower.endswith(ext):
            path += ext

        try:
            if fmt == "xlsx":
                self._write_xlsx(path)
            elif fmt == "csv":
                self._write_csv(path)
            else:
                with open(path, "w", encoding="utf-8") as f:
                    f.write(self._build_html())
            QMessageBox.information(self, "Export complete", f"Statistics exported to:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Export", f"Export failed: {e}")

    def _rows_for(self, result: TMResult):
        """Yield (label, segments, words, chars, tags, percent) for one TM,
        starting with the All total row."""
        total_words = result.total.words or 1
        yield ("All", result.total.segments, result.total.words,
               result.total.chars, result.total.tags, 100.0)
        for label in CATEGORIES:
            b = result.buckets[label]
            pct = b.words / total_words * 100.0
            yield (label, b.segments, b.words, b.chars, b.tags, pct)

    def _build_html(self) -> str:
        parts = [
            "<html><head><meta charset='utf-8'><style>",
            "body{font-family:Arial;font-size:12px;margin:24px;}",
            "h1{font-size:150%;} h2{font-size:120%;margin-top:1.5em;margin-bottom:.2em;}",
            ".scope{color:maroon;font-style:italic;margin-bottom:.5em;}",
            "table{border-collapse:collapse;width:100%;margin-bottom:1.5em;}",
            "th,td{border:1px solid #ccc;padding:4px 8px;text-align:right;}",
            "th:first-child,td:first-child{text-align:left;}",
            "tr.header{background:#b8c6cf;font-weight:bold;}",
            "tr.total td{font-weight:bold;background:#eef;}",
            "tr:nth-child(even){background:#f6f6ff;}",
            ".legend{font-size:11px;color:#444;margin-top:2em;}",
            ".legend dt{font-weight:bold;margin-top:.5em;}",
            ".legend dd{margin:0 0 0 1.5em;}",
            ".footer{margin-top:2em;padding-top:.8em;border-top:1px solid #ccc;"
            "font-size:11px;color:#666;}",
            "</style></head><body>",
            "<h1>Supervertaler Statistics</h1>",
        ]
        if self._project_name:
            parts.append(f"<p><b>Project:</b> {html.escape(self._project_name)}</p>")
        parts.append(
            f"<p>{len(self._segments):,} segments &middot; "
            f"{self._src_lang} &rarr; {self._tgt_lang}</p>"
        )

        for result in self._results:
            parts.append("<h2>Analysis</h2>")
            parts.append(f"<div class='scope'>Resources: {html.escape(result.tm_name)}</div>")
            parts.append("<table>")
            parts.append(
                "<tr class='header'><td>Type</td><td>Segments</td><td>Words</td>"
                "<td>Characters</td><td>Tags</td><td>Percent</td></tr>"
            )
            for label, segs, words, chars, tags, pct in self._rows_for(result):
                cls = " class='total'" if label == "All" else ""
                parts.append(
                    f"<tr{cls}><td>{html.escape(label)}</td>"
                    f"<td>{segs:,}</td><td>{words:,}</td>"
                    f"<td>{chars:,}</td><td>{tags:,}</td><td>{pct:.2f}</td></tr>"
                )
            parts.append("</table>")

        # Legend
        parts.append("<div class='legend'><b>What the match types mean</b><dl>")
        for label in CATEGORIES:
            parts.append(
                f"<dt>{html.escape(label)}</dt>"
                f"<dd>{html.escape(CATEGORY_HELP.get(label, ''))}</dd>"
            )
        parts.append("</dl></div>")

        # Footer link back to Supervertaler
        parts.append(
            "<div class='footer'>Generated by "
            "<a href='https://supervertaler.com/'>Supervertaler Workbench</a> "
            "&bull; by Michael Beijer</div>"
        )

        parts.append("</body></html>")
        return "\n".join(parts)

    def _write_csv(self, path: str):
        import csv
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Supervertaler Statistics"])
            if self._project_name:
                w.writerow(["Project", self._project_name])
            w.writerow([f"{len(self._segments)} segments",
                        f"{self._src_lang} -> {self._tgt_lang}"])
            w.writerow([])
            for result in self._results:
                w.writerow([f"Resources: {result.tm_name}"])
                w.writerow(["Type", "Segments", "Words", "Characters", "Tags", "Percent"])
                for label, segs, words, chars, tags, pct in self._rows_for(result):
                    w.writerow([label, segs, words, chars, tags, f"{pct:.2f}"])
                w.writerow([])
            # Legend
            w.writerow(["Legend"])
            for label in CATEGORIES:
                w.writerow([label, CATEGORY_HELP.get(label, "")])
            w.writerow([])
            w.writerow(["Generated by Supervertaler Workbench - https://supervertaler.com/"])

    def _write_xlsx(self, path: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistics"

        bold      = Font(bold=True)
        hdr_fill  = PatternFill("solid", fgColor="B8C6CF")
        total_fill = PatternFill("solid", fgColor="EEEEFF")

        r = 1
        c0 = ws.cell(row=r, column=1, value="Supervertaler Statistics")
        c0.font = Font(bold=True, size=14)
        r += 1
        if self._project_name:
            cell = ws.cell(row=r, column=1, value=f"Project: {self._project_name}")
            cell.font = bold
            r += 1
        ws.cell(row=r, column=1,
                value=f"{len(self._segments)} segments  ·  {self._src_lang} → {self._tgt_lang}")
        r += 2

        headers = ["Type", "Segments", "Words", "Characters", "Tags", "Percent"]
        for result in self._results:
            cell = ws.cell(row=r, column=1, value=f"Resources: {result.tm_name}")
            cell.font = bold
            r += 1
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=col, value=h)
                cell.font = bold
                cell.fill = hdr_fill
            r += 1
            for label, segs, words, chars, tags, pct in self._rows_for(result):
                ws.cell(row=r, column=1, value=label)
                ws.cell(row=r, column=2, value=segs)
                ws.cell(row=r, column=3, value=words)
                ws.cell(row=r, column=4, value=chars)
                ws.cell(row=r, column=5, value=tags)
                ws.cell(row=r, column=6, value=round(pct, 2))
                if label == "All":
                    for col in range(1, 7):
                        cell = ws.cell(row=r, column=col)
                        cell.font = bold
                        cell.fill = total_fill
                r += 1
            r += 1  # blank line between TMs

        # Legend
        cell = ws.cell(row=r, column=1, value="What the match types mean")
        cell.font = bold
        r += 1
        for label in CATEGORIES:
            ws.cell(row=r, column=1, value=label).font = bold
            ws.cell(row=r, column=2, value=CATEGORY_HELP.get(label, ""))
            r += 1
        r += 1
        ws.cell(row=r, column=1,
                value="Generated by Supervertaler Workbench — https://supervertaler.com/")

        # Reasonable column widths
        widths = [22, 12, 12, 14, 8, 10]
        for col, wdt in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = wdt

        wb.save(path)

    # ------------------------------------------------------------------

    def closeEvent(self, event):
        if self._worker and self._worker.isRunning():
            self._worker.cancel()
            self._worker.wait(3000)
        super().closeEvent(event)
